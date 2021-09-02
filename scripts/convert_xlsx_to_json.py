import datetime
from dateutil import parser
from dateutil.relativedelta import relativedelta
import json
from openpyxl import load_workbook
from pycel import ExcelCompiler
import string

PATH = r'D:\code\local\money-stuff\inc\Accounts.xlsx'
EXPORT = r'D:\code\local\money-stuff\inc\full-history.json'

SECTION_WIDTH = 6

ACCOUNT_ROWS = {
    'name': 3,
    'balance': 4,
    'venmo': 5,
    'check': 6,
    'transferBill': 7,
    'transferSav1': 8,
    'transferSav2': 9
}

TRANS_COLS_OFFSETS = {
    'name': 0,
    'value': 1,
    'value-pf': 2,
    'date': 3,
    'comment': 4
}

PAID_ACCOUNTS = (
    'PNC',
    'CapOneQS',
    'Discover',
    'CapOnePlat'
)

SAVING_ACCOUNTS = (
    'CIT',
    'Discover'
)

INITIAL_TRANS_ROW = 13
FINAL_TRANS_ROW = 33

MAPPER = {k:v for (k,v) in zip(range(1, len(string.ascii_uppercase) + 1), string.ascii_uppercase)}
MAPPER.update({k:'A'+v for (k,v) in zip(range(27, (len(string.ascii_uppercase) * 2) + 1), string.ascii_uppercase)})
MAPPER.update({v:k for (k,v) in MAPPER.items()})

starting_year = 2019
workbook = load_workbook(PATH)
compiler = ExcelCompiler(excel=workbook)
sheet_accounts = {}
sheet_trans = {}
for worksheet in workbook.worksheets:
    section_master_column = 1
    while section_master_column > 0:
        ############################################################
        # work out current active area and active date
        active_cell = worksheet.cell(1, section_master_column)
        if active_cell.is_date:
            active_date = active_cell.value
        else:
            if active_cell.data_type == 's':
                active_date = active_cell.value
                if not active_date or active_date.strip() == '':
                    section_master_column = -1
                    continue
            elif active_cell.data_type == 'n':
                if not active_cell.value:
                    section_master_column = -1
                    continue
            else:
                print(active_cell.data_type)
                section_master_column = -1
                continue

        if type(active_date) == str:
            active_date = parser.parse(active_date + ' ' + str(starting_year))

        # probably shouldn't just assume to increase the year, but it works
        if active_date.month == 1 and active_date.day == 1:
            starting_year = active_date.year + 1
            active_date = active_date + relativedelta(years=1)

        # print(active_date)

        ############################################################
        # processing

        # Scan for accounts
        accounts = []
        for i in range(0, 5):
            col = section_master_column + i
            account = {}
            for (name, row) in ACCOUNT_ROWS.items():
                cell = worksheet.cell(row, col)
                if not cell.value:
                    continue
                if type(cell.value) == str and cell.value.startswith('='):
                    account[name] = round(compiler.evaluate(f'{worksheet.title}!{cell.coordinate}'), 2)
                else:
                    account[name] = cell.value
            if not account or 'name' not in account:
                continue
            accounts.append(account)
        # print(accounts)

        # Scan for transactions
        trans = []
        for row in range(INITIAL_TRANS_ROW, FINAL_TRANS_ROW + 1):
            tran = {}
            for (name, i) in TRANS_COLS_OFFSETS.items():
                col = section_master_column + i
                cell = worksheet.cell(row, col)
                if not cell.value:
                    continue
                if type(cell.value) == str and cell.value.startswith('='):
                    tran[name] = round(compiler.evaluate(f'{worksheet.title}!{cell.coordinate}'), 2)
                elif name == 'comment' and type(cell.value) == str:
                    if cell.value == 'NA' or cell.value == 'N/A':
                        continue
                    elif cell.value.startswith('{') and cell.value.endswith('}'):
                        tran[name] = cell.value[1:-1]
                        continue
                    if cell.value.lower().split()[0].split('-')[0].strip() == 'paid':
                        tran['isPaid'] = True
                    if cell.value.lower().split()[0].split('-')[0].strip() == 'auto':
                        tran['auto'] = True
                    if cell.value in PAID_ACCOUNTS:
                        tran['isPaid'] = True
                        tran['paidFrom'] = cell.value
                    tran[name] = cell.value
                else:
                    tran[name] = cell.value
            if not tran:
                continue
            if not tran.get('name'):
                continue
            trans.append(tran)

        ############################################################
        # end processing
        
        sheet_accounts[active_date] = accounts
        sheet_trans[active_date] = trans

        section_master_column += SECTION_WIDTH

# begin creation of json object and export
def refval(ref, value):
    return {
        'ref': ref,
        'value': value
    }

def acc(account, sav2):
    name = account['name'] if account['name'] not in SAVING_ACCOUNTS else 'Savings'
    provider = account['name'] if account['name'] in SAVING_ACCOUNTS else 'Huntington'
    return {
        'name': name,
        'provider': provider,
        'balance': account.get('balance', 0),
        'income': list(filter(None, [
            refval('check', account['check']) if 'check' in account else None,
            refval('venmo' if account['name'] not in SAVING_ACCOUNTS else 'interest', account['venmo']) if 'venmo' in account else None
        ])),
        'transfer': list(filter(None, [
            refval('Huntington-BillPay', account.get('transferBill', 0)) if name != 'BillPay' else None,
            refval('Huntington-Savings', account.get('transferSav1', 0)) if name != 'Savings' else None,
            refval(f'{sav2}-Savings', account.get('transferSav2', 0)) if sav2 and name != sav2 else None,
        ])),
        'expenses': []
    }

def tra(tran, dt):
    if not tran.get('value', tran.get('value-pf')):
        return None
    return {
        'name': tran['name'],
        'value': tran['value-pf'] if tran.get('paidFrom') else tran.get('value', 0),
        'paidFrom': tran.get('paidFrom', None),
        'day': int(str(tran.get('date', dt.day)).replace("?", str(dt.day)).lower().replace("total", str(dt.day))),
        'isPaid': tran.get('isPaid', False),
        'auto': tran.get('auto', False),
        'comment': tran.get('comment', "")
    }

histories = []
templates = []

for dt in sheet_accounts:
    accounts = sheet_accounts[dt]
    transs = sheet_trans[dt]

    sav2 = None
    for account in accounts:
        if account['name'] == 'CIT':
            sav2 == 'CIT'
        if account['name'] == 'Discover':
            sav2 = 'Discover'

    accounts = [acc(account, sav2) for account in accounts]
    for account in accounts:
        if account['name'] == 'BillPay':
            break
        account = None
    if not account:
        continue
    account['expenses'] = list(filter(None, [tra(tran, dt) for tran in transs if tran]))

    now = datetime.datetime.now()
    if dt.date() <= now.date():
        histories.append({
            'date': dt.strftime('%Y-%m-%d'),
            'accounts': accounts
        })

with open(EXPORT, 'w') as f:
    json.dump({
        'templates': [],
        'history': histories
    }, f)

a = 20
